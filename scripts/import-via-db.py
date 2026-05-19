#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_AMBASSADORS = "/Users/Anselmo/Downloads/VIA Ambassador ed Expert - IWA&IWE - VIA Italian Wine Ambassadors - .csv"
DEFAULT_CANDIDATES = "/Users/Anselmo/Downloads/VIA 2026 CANDIDATES - MASTERSHEET DEF.xlsx"
SOURCE_VALUE = os.environ.get("VIA_DB_SOURCE_VALUE", "wine_awards")


def load_dotenv(path=".env"):
    env_path = Path(path)
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def clean(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"undefined", "null", "none", "nan"}:
        return None
    return text


def normalize_name(value):
    import unicodedata

    text = unicodedata.normalize("NFD", value or "")
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower()).strip()
    return text


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def is_missing(value):
    return not value or str(value).strip().lower() == "no data"


def slug(value):
    return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")


def year_from_course(value):
    match = re.search(r"(20\d{2})", value or "")
    return match.group(1) if match else None


def make_source_key(*parts):
    digest = hashlib.sha1("|".join(clean(part) or "" for part in parts).encode("utf-8")).hexdigest()[:18]
    return f"via_db:{digest}"


def location_city(city, state):
    city = clean(city)
    state = clean(state)
    if city and state and normalize_key(city) != normalize_key(state):
        return f"{city}, {state}"
    return city


def entry_identity(entry):
    if entry.get("email"):
        return f"email:{entry['email'].lower()}"
    if entry.get("phone"):
        return f"phone:{entry['normalized_name']}:{normalize_key(entry['phone'])}"
    return f"name-loc:{entry['normalized_name']}:{normalize_key(entry.get('city'))}:{normalize_key(entry.get('country'))}"


def parse_ambassadors(path):
    entries = []
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for index, row in enumerate(reader, start=2):
            first = clean(row.get("First Name"))
            last = clean(row.get("Last Name"))
            full = clean(row.get("Full Name")) or " ".join(part for part in [first, last] if part)
            if not full:
                continue
            course = clean(row.get("Course/class")) or clean(row.get("Course"))
            year = clean(row.get("Year")) or year_from_course(course)
            entry = {
                "first_name": first,
                "last_name": last,
                "full_name": full,
                "normalized_name": normalize_name(full),
                "email": clean(row.get("EMAIL")),
                "phone": clean(row.get("PHONE NUMBER")),
                "country": clean(row.get("Country / Region")) or "no data",
                "city": "no data",
                "state": None,
                "employer": None,
                "title": clean(row.get("JOB TITLE")),
                "occupation": clean(row.get("JOB TITLE")),
                "year": year,
                "course_class": course,
                "iwa_iwe": clean(row.get("IWA/IWE")),
                "source_file": Path(path).name,
                "source_sheet": None,
                "source_row": index,
                "raw": row,
            }
            entry["source_key"] = make_source_key(entry["source_file"], index, full, entry.get("email"), entry.get("phone"), year, course)
            entries.append(entry)
    return entries


def header_positions(headers):
    positions = {}
    for index, header in enumerate(headers):
        key = clean(header)
        if not key:
            continue
        positions.setdefault(key, []).append(index)
    return positions


def get_cell(row, positions, name, occurrence=0):
    indexes = positions.get(name) or []
    if len(indexes) <= occurrence:
        return None
    index = indexes[occurrence]
    return row[index] if index < len(row) else None


def parse_candidates(path):
    entries = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        positions = header_positions(headers)
        if "First Name" not in positions or "Last Name" not in positions:
            continue
        if "Email" not in positions and "Edition" not in positions:
            continue

        for row_index, row in enumerate(rows, start=2):
            first = clean(get_cell(row, positions, "First Name"))
            last = clean(get_cell(row, positions, "Last Name"))
            full = " ".join(part for part in [first, last] if part).strip()
            if not full:
                continue
            course = clean(get_cell(row, positions, "Edition")) or sheet.title
            city = location_city(
                get_cell(row, positions, "City", 1) or get_cell(row, positions, "City", 0),
                get_cell(row, positions, "State"),
            )
            country = clean(get_cell(row, positions, "Country")) or clean(get_cell(row, positions, "Nationality"))
            title = clean(get_cell(row, positions, "Job-Position")) or clean(get_cell(row, positions, "JOB TITLE"))
            raw = {
                clean(header) or f"column_{idx + 1}": clean(row[idx] if idx < len(row) else None)
                for idx, header in enumerate(headers)
            }
            entry = {
                "first_name": first,
                "last_name": last,
                "full_name": full,
                "normalized_name": normalize_name(full),
                "email": clean(get_cell(row, positions, "Email")) or clean(get_cell(row, positions, "EMAIL")),
                "phone": clean(get_cell(row, positions, "Phone")) or clean(get_cell(row, positions, "PHONE NUMBER")),
                "country": country or "no data",
                "city": city or "no data",
                "state": clean(get_cell(row, positions, "State")),
                "employer": clean(get_cell(row, positions, "Company")),
                "title": title,
                "occupation": title,
                "year": year_from_course(course),
                "course_class": course,
                "iwa_iwe": clean(get_cell(row, positions, "IWA/IWE")),
                "source_file": Path(path).name,
                "source_sheet": sheet.title,
                "source_row": row_index,
                "raw": raw,
            }
            entry["source_key"] = make_source_key(entry["source_file"], sheet.title, row_index, full, entry.get("email"), entry.get("phone"), course)
            entries.append(entry)
    return entries


def merge_people(entries):
    people = {}
    for entry in entries:
        key = entry_identity(entry)
        person = people.setdefault(key, {**entry, "sources": []})
        for field in ["email", "phone", "country", "city", "state", "employer", "title", "occupation", "year", "course_class", "iwa_iwe"]:
            if is_missing(person.get(field)) and not is_missing(entry.get(field)):
                person[field] = entry[field]
        person["sources"].append(entry)
    return list(people.values())


def rest_request(method, path, body=None, prefer=None):
    url = os.environ.get("VITE_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
      raise RuntimeError("Missing VITE_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env")
    data = None if body is None else json.dumps(body).encode("utf-8")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    request = urllib.request.Request(f"{url}{path}", data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request) as response:
            text = response.read().decode("utf-8")
            return json.loads(text) if text else None
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8")
        raise RuntimeError(f"Supabase {error.code}: {details}") from error


def fetch_all_contacts():
    contacts = []
    limit = 1000
    offset = 0
    while True:
        path = (
            "/rest/v1/contacts?select=id,full_name,first_name,last_name,normalized_name,email,city,country,"
            f"employer,title,occupation&limit={limit}&offset={offset}"
        )
        batch = rest_request("GET", path) or []
        contacts.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    return contacts


def contact_indexes(contacts):
    by_email = {}
    by_name_loc = {}
    by_name = {}
    for contact in contacts:
        email = clean(contact.get("email"))
        if email:
            by_email[email.lower()] = contact
        name = contact.get("normalized_name") or normalize_name(contact.get("full_name"))
        city = normalize_key(contact.get("city"))
        country = normalize_key(contact.get("country"))
        by_name_loc[f"{name}|{city}|{country}"] = contact
        by_name.setdefault(name, []).append(contact)
    return by_email, by_name_loc, by_name


def match_contact(person, indexes):
    by_email, by_name_loc, by_name = indexes
    if person.get("email") and person["email"].lower() in by_email:
        return by_email[person["email"].lower()], "email"
    key = f"{person['normalized_name']}|{normalize_key(person.get('city'))}|{normalize_key(person.get('country'))}"
    if not is_missing(person.get("city")) or not is_missing(person.get("country")):
        if key in by_name_loc:
            return by_name_loc[key], "name_location"
    matches = by_name.get(person["normalized_name"], [])
    if len(matches) == 1:
        return matches[0], "unique_name"
    return None, "new"


def contact_payload(person):
    return {
        "full_name": person["full_name"],
        "first_name": person.get("first_name"),
        "last_name": person.get("last_name"),
        "normalized_name": person["normalized_name"],
        "city": person.get("city") or "no data",
        "country": person.get("country") or "no data",
        "email": person.get("email"),
        "instagram_url": None,
        "linkedin_url": None,
        "employer": person.get("employer") or "no data",
        "title": person.get("title") or person.get("iwa_iwe") or "no data",
        "occupation": person.get("occupation") or person.get("iwa_iwe") or "no data",
        "status": "reviewed",
        "review_status": "unseen",
        "next_action": "da_verificare",
        "approval": False,
        "contacted": False,
        "notes": "VIA DB import.",
    }


def update_patch(contact, person):
    patch = {}
    for source_field, contact_field in [
        ("email", "email"),
        ("city", "city"),
        ("country", "country"),
        ("employer", "employer"),
        ("title", "title"),
        ("occupation", "occupation"),
    ]:
        value = person.get(source_field)
        current = contact.get(contact_field)
        if not is_missing(value) and is_missing(current):
            patch[contact_field] = value
    return patch


def upsert_contacts(people):
    rows = []
    for person in people:
        rows.append(contact_payload(person))
    returned = []
    for start in range(0, len(rows), 500):
        batch = rows[start:start + 500]
        query = urllib.parse.urlencode({"on_conflict": "normalized_name,country,city", "select": "id,normalized_name,country,city,email"})
        returned.extend(rest_request("POST", f"/rest/v1/contacts?{query}", batch, "resolution=merge-duplicates,return=representation") or [])
    return returned


def upsert_sources(source_rows):
    count = 0
    for start in range(0, len(source_rows), 500):
        batch = source_rows[start:start + 500]
        query = urllib.parse.urlencode({"on_conflict": "source,source_key", "select": "id"})
        rows = rest_request("POST", f"/rest/v1/contact_sources?{query}", batch, "resolution=merge-duplicates,return=representation") or []
        count += len(rows)
    return count


def build_source_rows(people):
    rows = []
    for person in people:
        contact_id = person.get("contact_id")
        if not contact_id:
            continue
        for source in person["sources"]:
            rows.append({
                "contact_id": contact_id,
                "source": SOURCE_VALUE,
                "source_key": source["source_key"],
                "restaurant_name": source.get("employer"),
                "award": " · ".join(part for part in [source.get("year"), source.get("iwa_iwe")] if part) or source.get("year"),
                "wine_role": source.get("course_class"),
                "profile_url": None,
                "raw_data": {
                    "source": "via_db",
                    "via_year": source.get("year"),
                    "via_course_class": source.get("course_class"),
                    "via_phone": source.get("phone"),
                    "via_iwa_iwe": source.get("iwa_iwe"),
                    "via_state": source.get("state"),
                    "via_file": source.get("source_file"),
                    "via_sheet": source.get("source_sheet"),
                    "via_row": source.get("source_row"),
                    "raw": source.get("raw"),
                },
            })
    return rows


def main():
    load_dotenv()
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambassadors", default=DEFAULT_AMBASSADORS)
    parser.add_argument("--candidates", default=DEFAULT_CANDIDATES)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    entries = parse_ambassadors(args.ambassadors) + parse_candidates(args.candidates)
    people = merge_people(entries)
    existing = fetch_all_contacts()
    indexes = contact_indexes(existing)

    matched = []
    new_people = []
    updates = []
    for person in people:
        contact, reason = match_contact(person, indexes)
        person["match_reason"] = reason
        if contact:
            person["contact_id"] = contact["id"]
            matched.append(person)
            patch = update_patch(contact, person)
            if patch:
                updates.append({"id": contact["id"], "full_name": contact["full_name"], "before": contact, "patch": patch})
        else:
            new_people.append(person)

    source_rows_preview = len(entries)
    preview = {
        "ambassadors_file": args.ambassadors,
        "candidates_file": args.candidates,
        "source_value_for_live_db": SOURCE_VALUE,
        "input_rows": len(entries),
        "unique_people_after_file_dedupe": len(people),
        "matched_existing_contacts": len(matched),
        "new_contacts": len(new_people),
        "contact_updates": len(updates),
        "source_rows": source_rows_preview,
        "sample_new": [contact_payload(person) for person in new_people[:5]],
        "sample_matches": [{"full_name": p["full_name"], "match_reason": p["match_reason"]} for p in matched[:10]],
    }

    Path("tmp").mkdir(exist_ok=True)
    Path("tmp/via-db-import-preview.json").write_text(json.dumps(preview, indent=2, ensure_ascii=False))
    Path("tmp/via-db-import-rollback.json").write_text(json.dumps({
        "created_at": dt.datetime.now(dt.timezone.utc).isoformat().replace("+00:00", "Z"),
        "matched_updates": updates,
    }, indent=2, ensure_ascii=False))

    print(json.dumps(preview, indent=2, ensure_ascii=False))

    if not args.apply:
        print("Dry run only. Re-run with --apply to update Supabase.")
        return

    for update in updates:
        rest_request("PATCH", f"/rest/v1/contacts?id=eq.{update['id']}", update["patch"])

    inserted = upsert_contacts(new_people)
    inserted_by_key = {
        f"{row['normalized_name']}|{normalize_key(row.get('city'))}|{normalize_key(row.get('country'))}": row["id"]
        for row in inserted
    }
    for person in new_people:
        key = f"{person['normalized_name']}|{normalize_key(person.get('city'))}|{normalize_key(person.get('country'))}"
        person["contact_id"] = inserted_by_key.get(key)

    missing_contact_ids = [person for person in people if not person.get("contact_id")]
    if missing_contact_ids:
        refreshed_indexes = contact_indexes(fetch_all_contacts())
        for person in missing_contact_ids:
            contact, _ = match_contact(person, refreshed_indexes)
            if contact:
                person["contact_id"] = contact["id"]

    source_rows = build_source_rows(people)
    source_count = upsert_sources(source_rows)

    result = {
        **preview,
        "applied": True,
        "updated_existing_contacts": len(updates),
        "upserted_new_contacts": len(inserted),
        "source_rows_built": len(source_rows),
        "source_rows_missing_contact": len(people) - sum(1 for person in people if person.get("contact_id")),
        "upserted_source_rows": source_count,
    }
    Path("tmp/via-db-import-result.json").write_text(json.dumps(result, indent=2, ensure_ascii=False))
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(exc, file=sys.stderr)
        sys.exit(1)
