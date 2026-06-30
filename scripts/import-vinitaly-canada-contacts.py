#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DB_SOURCE_VALUE = "wine_awards"
NO_DATA = "no data"

DEFAULT_FILES = [
    {
        "source": "elenco_buyer",
        "event": "Elenco Buyer",
        "path": "/Users/Anselmo/Downloads/Elenco Buyer - Elenco.csv",
        "kind": "csv",
        "columns": {
            "company": "Company Name",
            "first": "Name",
            "last": "Surname",
            "title": "Job Title",
            "category": "AGENZIA",
            "email": "E-mail address",
            "phone": "Mobile phone number\n (with international dial code)",
            "website": "Web page",
            "city": "City",
            "country": "Country",
        },
    },
    {
        "source": "rsvp_spazio_camera_vinitaly_canada",
        "event": "22 GIUGNO RSVP SPAZIO CAMERA VINITALY CANADA",
        "path": "/Users/Anselmo/Downloads/22 GIUGNO_RSVP SPAZIO CAMERA_VINITALY CANADA.xlsx",
        "kind": "xlsx",
        "columns": {
            "company": "AZIENDA",
            "first": "NOME",
            "last": "COGNOME",
            "title": "TITOLO",
            "category": "CATEGORY",
            "email": "EMAIL",
        },
    },
    {
        "source": "rsvp_cena_aria_vinitaly_canada",
        "event": "23 GIUGNO RSVP CENA ARIA VINITALY CANADA",
        "path": "/Users/Anselmo/Downloads/23 GIUGNO_RSVP CENA ARIA_VINITALY CANADA.xlsx",
        "kind": "xlsx",
        "columns": {
            "company": "COMPANY",
            "first": "FIRST NAME",
            "last": "LAST NAME",
            "title": "TITLE",
            "category": "CATEGORY",
            "email": "EMAIL",
        },
    },
    {
        "source": "rsvp_ambasciata_vinitaly_canada",
        "event": "24 GIUGNO RSVP AMBASCIATA VINITALY CANADA",
        "path": "/Users/Anselmo/Downloads/24 GIUGNO_RSVP AMBASCIATA_VINITALY CANADA.xlsx",
        "kind": "xlsx",
        "columns": {
            "company": "COMPANY",
            "first": "First Name",
            "last": "Last Name",
            "title": "Title",
            "category": "Category",
            "email": "EMAIL",
        },
    },
    {
        "source": "rsvp_via_course_vinitaly_canada",
        "event": "22-23 GIUGNO RSVP VIA COURSE VINITALY CANADA",
        "path": "/Users/Anselmo/Downloads/22-23 GIUGNO_RSVP VIA COURSE_VINITALY CANADA.xlsx",
        "kind": "xlsx",
        "columns": {
            "company": "Company",
            "first": "Primo Nome",
            "last": "Cognome",
            "title": "Job Title",
            "category": "Industry",
            "email": "Email",
            "phone": "Phone Number",
            "address": "Address",
        },
    },
]


def load_dotenv(path=".env"):
    env_path = Path(path)
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def clean(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        value = value.date().isoformat()
    elif isinstance(value, dt.date):
        value = value.isoformat()
    elif isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"undefined", "null", "none", "nan"}:
        return None
    return re.sub(r"\s+", " ", text)


def is_missing(value):
    return not clean(value) or clean(value).lower() == NO_DATA


def normalize_name(value):
    import unicodedata

    text = unicodedata.normalize("NFD", value or "")
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-zA-Z0-9]+", " ", text.lower()).strip()


def normalize_email(value):
    text = clean(value)
    if not text:
        return None
    match = re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, flags=re.IGNORECASE)
    return match.group(0).lower() if match else None


def extract_emails(value):
    text = clean(value) or ""
    return [email.lower() for email in re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", text, flags=re.IGNORECASE)]


def normalize_url(value):
    text = clean(value)
    if not text:
        return None
    if "@" in text and not text.lower().startswith(("http://", "https://", "www.")):
        return None
    if text.lower().startswith(("http://", "https://")):
        return text
    if "." in text and " " not in text:
        return f"https://{text}"
    return None


def source_key(source, *parts):
    digest = hashlib.sha1("|".join(clean(part) or "" for part in parts).encode("utf-8")).hexdigest()[:18]
    return f"{source}:{digest}"


def row_value(row, columns, field):
    column = columns.get(field)
    return clean(row.get(column)) if column else None


def read_csv(config):
    path = Path(config["path"])
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        for index, row in enumerate(csv.DictReader(handle, dialect=dialect), start=2):
            yield None, index, row


def read_xlsx(config):
    path = Path(config["path"])
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean(value) or f"column_{idx + 1}" for idx, value in enumerate(next(rows))]
        except StopIteration:
            continue
        for row_index, values in enumerate(rows, start=2):
            row = {header: clean(values[idx] if idx < len(values) else None) for idx, header in enumerate(headers)}
            yield sheet.title, row_index, row


def build_entry(config, sheet, row_index, row):
    columns = config["columns"]
    first = row_value(row, columns, "first")
    last = row_value(row, columns, "last")
    full_name = " ".join(part for part in [first, last] if part).strip()
    company = row_value(row, columns, "company")
    title = row_value(row, columns, "title")
    category = row_value(row, columns, "category") or NO_DATA
    email_raw = row_value(row, columns, "email")
    emails = extract_emails(email_raw)
    email = emails[0] if emails else None
    phone = row_value(row, columns, "phone")
    website = normalize_url(row_value(row, columns, "website"))
    address = row_value(row, columns, "address")
    city = row_value(row, columns, "city") or NO_DATA
    country = row_value(row, columns, "country") or NO_DATA

    if not full_name and not email and not company:
        return None
    if normalize_name(full_name) == "volunteer via" and not email and not company:
        return None

    instagram_url = website if website and "instagram.com" in website.lower() else None
    linkedin_url = website if website and "linkedin.com" in website.lower() else None
    profile_url = website if website and not instagram_url and not linkedin_url else None

    if not full_name:
        full_name = email or company

    raw = {key: clean(value) for key, value in row.items() if clean(value)}
    raw.update({
        "source": config["source"],
        "source_file": Path(config["path"]).name,
        "source_sheet": sheet,
        "source_row": row_index,
        "event": config["event"],
        "category": category,
        "phone": phone,
        "website": website,
        "address": address,
        "emails": emails,
    })

    key = source_key(config["source"], Path(config["path"]).name, sheet, row_index, full_name, email, company)
    return {
        "source": config["source"],
        "source_key": key,
        "source_file": Path(config["path"]).name,
        "source_sheet": sheet,
        "source_row": row_index,
        "event": config["event"],
        "first_name": first,
        "last_name": last,
        "full_name": full_name,
        "normalized_name": normalize_name(full_name),
        "email": email,
        "email_raw": email_raw,
        "instagram_url": instagram_url,
        "linkedin_url": linkedin_url,
        "profile_url": profile_url,
        "employer": company,
        "title": title,
        "occupation": title,
        "city": city,
        "country": country,
        "category": category,
        "phone": phone,
        "website": website,
        "address": address,
        "raw": raw,
    }


def load_entries(configs):
    entries = []
    skipped = 0
    for config in configs:
        path = Path(config["path"])
        if not path.exists():
            raise FileNotFoundError(path)
        reader = read_csv(config) if config["kind"] == "csv" else read_xlsx(config)
        for sheet, row_index, row in reader:
            entry = build_entry(config, sheet, row_index, row)
            if entry:
                entries.append(entry)
            else:
                skipped += 1
    return entries, skipped


def rest_request(method, path, body=None, prefer=None):
    url = os.environ.get("VITE_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("Missing VITE_SUPABASE_URL/SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")
    data = None if body is None else json.dumps(body).encode("utf-8")
    request = urllib.request.Request(f"{url}{path}", data=data, method=method)
    request.add_header("apikey", key)
    request.add_header("Authorization", f"Bearer {key}")
    request.add_header("Content-Type", "application/json")
    if prefer:
        request.add_header("Prefer", prefer)
    with urllib.request.urlopen(request) as response:
        text = response.read().decode("utf-8")
        return json.loads(text) if text else None


def check_supabase_connection():
    rest_request("GET", "/rest/v1/contacts?select=id&limit=1")


def quote(value):
    return urllib.parse.quote(str(value), safe="")


def fetch_existing(entry):
    if entry.get("email"):
        rows = rest_request("GET", f"/rest/v1/contacts?select=*&email=eq.{quote(entry['email'])}&limit=1") or []
        if rows:
            return rows[0]

    rows = rest_request(
        "GET",
        "/rest/v1/contacts?select=*&"
        f"normalized_name=eq.{quote(entry['normalized_name'])}&"
        f"country=eq.{quote(entry['country'])}&"
        f"city=eq.{quote(entry['city'])}&limit=1",
    ) or []
    if rows:
        return rows[0]

    if entry.get("employer"):
        rows = rest_request(
            "GET",
            "/rest/v1/contacts?select=*&"
            f"normalized_name=eq.{quote(entry['normalized_name'])}&"
            f"employer=eq.{quote(entry['employer'])}&limit=1",
        ) or []
        if rows:
            return rows[0]
    return None


def contact_insert_payload(entry):
    ready = bool(entry.get("email") or entry.get("instagram_url") or entry.get("linkedin_url"))
    return {
        "full_name": entry["full_name"],
        "first_name": entry.get("first_name"),
        "last_name": entry.get("last_name"),
        "normalized_name": entry["normalized_name"],
        "email": entry.get("email"),
        "instagram_url": entry.get("instagram_url"),
        "linkedin_url": entry.get("linkedin_url"),
        "employer": entry.get("employer"),
        "title": entry.get("title"),
        "occupation": entry.get("occupation"),
        "city": entry.get("city") or NO_DATA,
        "country": entry.get("country") or NO_DATA,
        "notes": None,
        "status": "todo",
        "review_status": "unseen",
        "next_action": "pronto_da_contattare" if ready else None,
        "approval": False,
        "contacted": False,
        "assigned_to": None,
    }


def contact_patch_payload(existing, entry):
    patch = {}
    for field in ["first_name", "last_name", "email", "instagram_url", "linkedin_url", "employer", "title", "occupation", "city", "country"]:
        value = entry.get(field)
        if value and is_missing(existing.get(field)):
            patch[field] = value
    if entry.get("email") and is_missing(existing.get("next_action")):
        patch["next_action"] = "pronto_da_contattare"
    return patch


def upsert_contact(entry, apply, offline=False):
    if offline:
        return None, "insert"
    existing = fetch_existing(entry)
    if not apply:
        return existing, "update" if existing else "insert"

    if existing:
        patch = contact_patch_payload(existing, entry)
        if patch:
            rows = rest_request(
                "PATCH",
                f"/rest/v1/contacts?id=eq.{quote(existing['id'])}&select=*",
                patch,
                "return=representation",
            ) or []
            existing = rows[0] if rows else existing
        return existing, "update"

    try:
        rows = rest_request("POST", "/rest/v1/contacts?select=*", contact_insert_payload(entry), "return=representation") or []
        return rows[0], "insert"
    except Exception:
        existing = fetch_existing(entry)
        if not existing:
            raise
        return existing, "update"


def upsert_source(contact, entry, apply):
    row = {
        "contact_id": contact["id"],
        "source": DB_SOURCE_VALUE,
        "source_key": entry["source_key"],
        "restaurant_name": entry.get("employer"),
        "award": entry.get("event"),
        "wine_role": entry.get("category"),
        "profile_url": entry.get("profile_url"),
        "raw_data": entry.get("raw"),
    }
    if apply:
        rest_request(
            "POST",
            "/rest/v1/contact_sources?on_conflict=source,source_key",
            row,
            "resolution=merge-duplicates",
        )


def summarize(entries, skipped):
    by_source = {}
    by_category = {}
    for entry in entries:
        by_source[entry["source"]] = by_source.get(entry["source"], 0) + 1
        by_category.setdefault(entry["source"], {})
        category = entry.get("category") or NO_DATA
        by_category[entry["source"]][category] = by_category[entry["source"]].get(category, 0) + 1
    return {
        "entries": len(entries),
        "skipped_blank_or_placeholder": skipped,
        "with_email": sum(1 for entry in entries if entry.get("email")),
        "with_phone": sum(1 for entry in entries if entry.get("phone")),
        "by_source": by_source,
        "categories_by_source": by_category,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Write contacts and sources to Supabase")
    parser.add_argument("--env", default=".env")
    parser.add_argument("--offline", action="store_true", help="Only parse files and summarize without querying Supabase")
    args = parser.parse_args()

    load_dotenv(args.env)
    entries, skipped = load_entries(DEFAULT_FILES)
    summary = summarize(entries, skipped)
    print(json.dumps({"dry_run": not args.apply, **summary}, ensure_ascii=False, indent=2))

    if args.offline:
      print(json.dumps({
          "applied": False,
          "offline": True,
          "contacts_parsed": len(entries),
          "source_rows_parsed": len(entries),
          "db_source_value": DB_SOURCE_VALUE,
      }, ensure_ascii=False, indent=2))
      return

    try:
        check_supabase_connection()
    except Exception as exc:
        print(json.dumps({
            "applied": False,
            "error": "Supabase non raggiungibile da questa sessione. Riesegui il comando da un ambiente con rete/DNS abilitati.",
            "details": str(exc),
            "command": "python3 scripts/import-vinitaly-canada-contacts.py --apply",
        }, ensure_ascii=False, indent=2), file=sys.stderr)
        raise SystemExit(1)

    inserted = updated = source_rows = errors = 0
    for entry in entries:
        try:
            contact, action = upsert_contact(entry, args.apply)
            if action == "insert":
                inserted += 1
            else:
                updated += 1
            if contact:
                upsert_source(contact, entry, args.apply)
                source_rows += 1
        except Exception as exc:
            errors += 1
            print(f"ERROR {entry['source_file']} row {entry['source_row']} {entry['full_name']}: {exc}", file=sys.stderr)

    print(json.dumps({
        "applied": args.apply,
        "contacts_to_insert" if not args.apply else "contacts_inserted": inserted,
        "contacts_to_update" if not args.apply else "contacts_updated_or_matched": updated,
        "source_rows_to_upsert" if not args.apply else "source_rows_upserted": source_rows,
        "errors": errors,
        "db_source_value": DB_SOURCE_VALUE,
    }, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
